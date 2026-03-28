# core/excel_report.py - НОВАЯ СТРУКТУРА ОТЧЁТА
import pandas as pd
from datetime import datetime
import os
from collections import Counter

class ExcelReport:
    """Создание отчёта с новой структурой: 3 листа"""
    
    def __init__(self, db, data_dir='data'):
        self.db = db
        self.data_dir = data_dir
        self.sources_file = os.path.join(data_dir, 'sources.json')
    
    def _determine_region_flags(self, info):
        """Определение региональных флагов"""
        ru = info.get('ru_access', False)
        us = info.get('us_access', False)
        country_code = info.get('country_code', '')
        region = info.get('region', '')
        
        if not ru and country_code == 'RU':
            ru = True
        if not ru and region == 'ru':
            ru = True
        if not us and country_code == 'US':
            us = True
        if not us and region == 'us':
            us = True
        
        return ru, us
    
    def _get_country_from_info(self, info):
        """Получение названия страны"""
        country_code = info.get('country_code', '')
        country = info.get('country', '')
        
        if country:
            return country
        elif country_code:
            # Преобразуем код страны в название
            countries = {
                'RU': 'Россия', 'US': 'США', 'GB': 'Великобритания',
                'DE': 'Германия', 'FR': 'Франция', 'NL': 'Нидерланды',
                'CA': 'Канада', 'AU': 'Австралия', 'JP': 'Япония',
                'CN': 'Китай', 'SG': 'Сингапур', 'IN': 'Индия',
                'BR': 'Бразилия', 'KR': 'Корея', 'HK': 'Гонконг',
                'VN': 'Вьетнам', 'ID': 'Индонезия', 'PH': 'Филиппины',
                'EC': 'Эквадор', 'CO': 'Колумбия', 'DO': 'Доминикана',
                'DE': 'Германия', 'BG': 'Болгария', 'CZ': 'Чехия'
            }
            return countries.get(country_code, country_code)
        return 'Неизвестно'
    
    def create_report(self, filename='reports/proxy_report.xlsx'):
        """Создание отчёта с новой структурой"""
        print("📊 Создаю Excel-отчёт (новая структура)...")
        
        os.makedirs(os.path.dirname(filename) if os.path.dirname(filename) else '.', exist_ok=True)
        
        # Собираем данные
        proxies_dict = self.db.db.get('proxies', {})
        stats = self.db.get_stats()
        
        # ========== ЛИСТ 1: Прокси с распределением ==========
        data_all = []
        data_global = []
        data_russian = []
        data_american = []
        data_other = []  # Остальные с указанием страны
        
        for proxy, info in proxies_dict.items():
            if not info.get('working'):
                continue
            
            ru, us = self._determine_region_flags(info)
            country = self._get_country_from_info(info)
            latency = info.get('latency', 9999)
            
            row = {
                'Прокси': proxy,
                'Задержка (мс)': latency if latency != 9999 else '—',
                'Регион': '🌍 Глобальный' if (ru and us) else '🇷🇺 Россия' if ru else '🇺🇸 США' if us else '🌎 Другой',
                'Страна': country if not (ru or us) else '',
                'Дата добавления': self._format_date(info.get('first_seen')),
                'Источник': info.get('source', 'неизвестен')
            }
            data_all.append(row)
            
            # Распределение по категориям
            if ru and us:
                data_global.append({'Прокси': proxy, 'Задержка (мс)': latency})
            elif ru:
                data_russian.append({'Прокси': proxy, 'Задержка (мс)': latency})
            elif us:
                data_american.append({'Прокси': proxy, 'Задержка (мс)': latency})
            else:
                data_other.append({'Прокси': proxy, 'Задержка (мс)': latency, 'Страна': country})
        
        # Создаём Excel файл с новыми листами
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            
            # --- ЛИСТ 1: Все прокси с распределением ---
            df_all = pd.DataFrame(data_all)
            if not df_all.empty:
                df_all = df_all.sort_values('Задержка (мс)', ascending=True)
                df_all.to_excel(writer, sheet_name='Все прокси', index=False)
            else:
                pd.DataFrame({'Сообщение': ['Нет данных']}).to_excel(writer, sheet_name='Все прокси', index=False)
            
            # --- ЛИСТ 2: Статистика ---
            # Подсчёт по странам для "Остальных"
            other_countries = Counter()
            for proxy, info in proxies_dict.items():
                if info.get('working'):
                    ru, us = self._determine_region_flags(info)
                    if not ru and not us:
                        country = self._get_country_from_info(info)
                        other_countries[country] += 1
            
            stats_data = [
                ['📊 ОБЩАЯ СТАТИСТИКА', ''],
                ['Всего прокси в базе', stats['total_in_db']],
                ['Рабочих прокси', stats['working_now']],
                ['Нерабочих прокси', stats['total_in_db'] - stats['working_now']],
                ['', ''],
                ['📍 ГЕОГРАФИЧЕСКОЕ РАСПРЕДЕЛЕНИЕ', ''],
                ['Глобальные (РФ+США)', stats['global']],
                ['Российские (только РФ)', stats['russian']],
                ['Американские (только США)', stats['american']],
                ['Остальные страны', sum(other_countries.values())],
                ['', ''],
                ['🌍 РАСПРЕДЕЛЕНИЕ ПО СТРАНАМ (ОСТАЛЬНЫЕ)', ''],
            ]
            
            # Добавляем страны
            for country, count in sorted(other_countries.items(), key=lambda x: -x[1]):
                stats_data.append([f'  {country}', count])
            
            stats_data.extend([
                ['', ''],
                ['🕐 ИНФОРМАЦИЯ ОБ ОБНОВЛЕНИИ', ''],
                ['Последнее обновление базы', self._format_date(self.db.db['stats'].get('last_update'))],
                ['Всего проверено за всё время', stats['total_seen']],
                ['Дата создания отчёта', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ])
            
            df_stats = pd.DataFrame(stats_data, columns=['Показатель', 'Значение'])
            df_stats.to_excel(writer, sheet_name='Статистика', index=False)
            
            # --- ЛИСТ 3: Источники прокси ---
            sources_data = []
            sources_set = set()
            
            for proxy, info in proxies_dict.items():
                source = info.get('source', 'неизвестен')
                if source not in sources_set and source != 'неизвестен':
                    sources_set.add(source)
                    # Пытаемся извлечь дату первого появления источника
                    first_seen = info.get('first_seen', '')
                    sources_data.append({
                        'Источник': source,
                        'Дата первого обнаружения': self._format_date(first_seen),
                        'Количество прокси': sum(1 for p, i in proxies_dict.items() if i.get('source') == source and i.get('working')),
                        'Статус': 'Активен'
                    })
            
            # Добавляем стандартные источники из конфигурации
            known_sources = [
                'thespeedx_http', 'thespeedx_socks4', 'thespeedx_socks5',
                'jetkai_http', 'jetkai_socks4', 'jetkai_socks5',
                'proxyscrape_api', 'proxymania', 'ru_scrape', 'us_scrape'
            ]
            
            for src in known_sources:
                if src not in sources_set:
                    sources_data.append({
                        'Источник': src,
                        'Дата первого обнаружения': '-',
                        'Количество прокси': 0,
                        'Статус': 'Ожидает'
                    })
            
            df_sources = pd.DataFrame(sources_data)
            if not df_sources.empty:
                df_sources = df_sources.sort_values('Количество прокси', ascending=False)
                df_sources.to_excel(writer, sheet_name='Источники', index=False)
            else:
                pd.DataFrame({'Сообщение': ['Нет данных об источниках']}).to_excel(writer, sheet_name='Источники', index=False)
        
        # Форматирование Excel
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        
        wb = load_workbook(filename)
        
        # Форматируем каждый лист
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Автоширина колонок
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Заголовки жирным
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
        
        wb.save(filename)
        print(f"✅ Отчёт сохранён: {filename}")
        return filename
    
    def _format_date(self, date_str):
        if not date_str:
            return '-'
        try:
            dt = datetime.fromisoformat(date_str)
            return dt.strftime('%Y-%m-%d %H:%M:%S')
        except:
            return date_str[:19] if len(date_str) > 19 else date_str
